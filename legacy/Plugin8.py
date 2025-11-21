import os
from openpyxl import load_workbook
from datetime import datetime

SRC_DIR = r"C:\Users\WerlandL\FRET3\XLSX"

files = [f for f in os.listdir(SRC_DIR) if f.lower().startswith("ed") and f.lower().endswith(".xlsx")]
print(f"Found {len(files)} .xlsx files in {SRC_DIR!r}")

for fname in files:
    path = os.path.join(SRC_DIR, fname)
    print(f"\n→ Processing {fname}")

    try:
        wb_val = load_workbook(path, data_only=True)
        ws_val = wb_val.active
        wb     = load_workbook(path)
        ws     = wb.active

        # 1) find last used column in row 56
        max_col = ws.max_column
        for c in range(max_col, 0, -1):
            if ws.cell(row=56, column=c).value is not None:
                max_col = c
                break

        # 2) Paste-special: copy computed values from rows 56–72 to 74–90
        for col in range(2, max_col+1, 2):  # B=2, D=4, F=6, ...
            for r_src, r_dst in zip(range(56, 73), range(74, 91)):
                val = ws_val.cell(row=r_src, column=col).value
                ws.cell(row=r_dst, column=col, value=val)

        # 3) (re)create TransposedData sheet
        if "TransposedData" in wb.sheetnames:
            wb.remove(wb["TransposedData"])
        ws2 = wb.create_sheet("TransposedData")

        # 4) write header: A1="Channel", B1–R1 = original A56:A72
        ws2.cell(row=1, column=1, value="Channel")
        headers = [ws.cell(row=r, column=1).value for r in range(56, 73)]
        for j, h in enumerate(headers, start=2):
            ws2.cell(row=1, column=j, value=h)

        # 5) write transposed data rows with new labels
        dest_row = 2
        file_id = fname[3:].split('.')[0]  # strip 'Ed_' prefix and extension

        for col in range(2, max_col+1, 2):
            channel = ws.cell(row=1, column=col).value
            series  = [ws.cell(r, column=col).value for r in range(74, 91)]
            if all(v is None for v in series):
                continue
            new_label = f"{file_id}_{channel}"
            ws2.cell(row=dest_row, column=1, value=new_label)
            for idx, v in enumerate(series, start=2):
                ws2.cell(row=dest_row, column=idx, value=v)
            dest_row += 1

        # 6) delete the very last row of TransposedData
        last = ws2.max_row
        ws2.delete_rows(last, 1)

        # 7) save and close
        wb.save(path)
        mtime = os.path.getmtime(path)
        print(f"   ✔ TransposedData created, saved at {datetime.fromtimestamp(mtime)}")

        wb.close()
        wb_val.close()

    except Exception as e:
        print(f"   ⚠ Error with {fname}: {e}")

print("\nAll files processed.")
