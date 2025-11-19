import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from statistics import mean, stdev

SRC_DIR = r"C:\Users\WerlandL\FRET3\XLSX"

def get_matching_files(pattern):
    return sorted([
        f for f in os.listdir(SRC_DIR)
        if f.lower().endswith(".xlsx")
        and "ed_" in f.lower()
        and pattern.lower() in f.lower()
    ])

def format_value(val):
    if isinstance(val, float):
        return round(val, 3)  # Keep 3-digit precision internally
    return val

def process_group(file_list, output_name):
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Day Summary"
    row_out = 1

    bold_font = Font(bold=True)
    italic_font = Font(italic=True)
    bold_keys = ["Channel", "Area", "Major", "Minor", "DonBB", "DonAB", "AccBB", "AccAB",
                 "BG don", "BG acc", "FRET", "Bleached", "CorrF", "DonBB-BG", "DonAB-BG", "CorrFRET",
                 "AccBB-BG", "ACCAB-BG", "Fac"]

    apply_bleached_filter = output_name.lower().endswith("in")
    italic_corrfret = output_name.lower().endswith("out")

    for fname in file_list:
        path = os.path.join(SRC_DIR, fname)
        try:
            wb = load_workbook(path, data_only=True)
            if "TransposedData" not in wb.sheetnames:
                continue
            ws = wb["TransposedData"]
            data = list(ws.iter_rows(values_only=True))
            headers = list(data[0])
            rows = data[1:]

            try:
                bleached_idx = headers.index("Bleached")
            except ValueError:
                bleached_idx = -1
            try:
                corrfret_idx = headers.index("CorrFRET")
            except ValueError:
                corrfret_idx = -1
            try:
                donab_bg_idx = headers.index("DonAB-BG")
                accbb_bg_idx = headers.index("AccBB-BG")
            except ValueError:
                print(f"⚠️ Missing DonAB-BG or AccBB-BG in {fname}, skipping.")
                continue

            if apply_bleached_filter and bleached_idx >= 0:
                rows = [
                    r for r in rows
                    if isinstance(r[bleached_idx], (int, float)) and r[bleached_idx] >= 69.6
                ]

            headers.append("Ratio")

            # Write header
            for col_idx, val in enumerate(headers, start=1):
                cell = ws_out.cell(row=row_out, column=col_idx, value=val)
                cell.font = bold_font
                if italic_corrfret and col_idx - 1 == corrfret_idx:
                    cell.font = Font(bold=True, italic=True)
            row_out += 1

            ratio_values = []

            # Write data rows
            for row in rows:
                new_row = list(row)
                val1 = row[accbb_bg_idx]
                val2 = row[donab_bg_idx]
                ratio = None
                if isinstance(val1, (int, float)) and isinstance(val2, (int, float)) and val2 != 0:
                    ratio = val1 / val2
                    ratio_values.append(ratio)
                new_row.append(ratio)

                bold_row = str(row[0]).strip() in bold_keys
                for col_idx, val in enumerate(new_row, start=1):
                    val = format_value(val)
                    cell = ws_out.cell(row=row_out, column=col_idx, value=val)
                    if isinstance(val, float):
                        cell.number_format = '0.0'  # Show only 1 digit in Excel
                    if bold_row:
                        cell.font = bold_font
                    if italic_corrfret and col_idx - 1 == corrfret_idx:
                        cell.font = Font(italic=True)
                row_out += 1

            row_out += 1

            # Write averages
            ws_out.cell(row=row_out, column=1, value="Average").font = bold_font
            for col_idx in range(2, len(headers)):
                values = [r[col_idx - 1] for r in rows if isinstance(r[col_idx - 1], (int, float))]
                if values:
                    val = format_value(mean(values))
                    cell = ws_out.cell(row=row_out, column=col_idx, value=val)
                    if isinstance(val, float):
                        cell.number_format = '0.0'
                    cell.font = bold_font
                    if italic_corrfret and col_idx - 1 == corrfret_idx:
                        cell.font = Font(bold=True, italic=True)
            if ratio_values:
                val = format_value(mean(ratio_values))
                cell = ws_out.cell(row=row_out, column=len(headers), value=val)
                if isinstance(val, float):
                    cell.number_format = '0.0'
                cell.font = bold_font
            row_out += 1

            # Write stdevs
            ws_out.cell(row=row_out, column=1, value="Stdev").font = bold_font
            for col_idx in range(2, len(headers)):
                values = [r[col_idx - 1] for r in rows if isinstance(r[col_idx - 1], (int, float))]
                if len(values) > 1:
                    val = format_value(stdev(values))
                    cell = ws_out.cell(row=row_out, column=col_idx, value=val)
                    if isinstance(val, float):
                        cell.number_format = '0.0'
                    cell.font = bold_font
                    if italic_corrfret and col_idx - 1 == corrfret_idx:
                        cell.font = Font(bold=True, italic=True)
            if len(ratio_values) > 1:
                val = format_value(stdev(ratio_values))
                cell = ws_out.cell(row=row_out, column=len(headers), value=val)
                if isinstance(val, float):
                    cell.number_format = '0.0'
                cell.font = bold_font
            row_out += 2

        except Exception as e:
            print(f"⚠️ Error with file {fname}: {e}")

    out_path = os.path.join(SRC_DIR, f"{output_name}.xlsx")
    wb_out.save(out_path)
    print(f"✅ Saved {output_name}.xlsx with {len(file_list)} datasets.")

# Run for both groups
in_files = get_matching_files("IN-Ble")
out_files = get_matching_files("OUT")

process_group(in_files, "DaySummaryIN")
process_group(out_files, "DaySummaryOUT")
